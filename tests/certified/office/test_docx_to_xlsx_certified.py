from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from app.core.settings import settings
from app.main import app


def test_docx_to_xlsx_conversion_creates_valid_xlsx(tmp_path: Path):
    client = TestClient(app)

    docx_path = Path("tests/assets/regression/sample.docx")
    assert docx_path.exists(), "Sample DOCX is missing"

    resp = client.post(
        "/convert",
        data={"target_format": "xlsx", "operation": "docx-to-xlsx"},
        files={
            "file": (
                docx_path.name,
                docx_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )

    assert resp.status_code == 201, resp.text
    payload = resp.json()
    assert payload.get("status") == "success"
    download_path = payload.get("download_path")
    assert download_path, payload
    assert download_path.startswith("/download/")
    relative_parts = Path(download_path.removeprefix("/download/")).parts
    assert len(relative_parts) == 2, f"Unexpected download path shape: {download_path}"
    conversion_id, filename = relative_parts
    assert filename.endswith(".xlsx")

    local_path = settings.OUTPUT_DIR / conversion_id / filename
    download_resp = client.get(download_path)
    assert download_resp.status_code == 200, download_resp.text
    assert download_resp.content, "Downloaded content is empty"

    assert local_path.exists(), f"Expected output XLSX not found: {local_path}"
    assert local_path.stat().st_size > 0, "Output XLSX is empty"
    assert local_path.suffix.lower() == ".xlsx"

    # Content-integrity: re-open output with openpyxl and verify the source
    # DOCX text actually made it into the workbook (not a fake file).
    workbook = openpyxl.load_workbook(str(local_path), data_only=True, read_only=True)
    try:
        assert "Content" in workbook.sheetnames, f"Missing Content sheet: {workbook.sheetnames}"

        sheet = workbook["Content"]
        values = [cell[0] for row in sheet.iter_rows(values_only=True) for cell in [row]]
        joined = "\n".join(str(v) for v in values if v is not None)
        assert "Sample DOCX" in joined, f"Expected source paragraph in XLSX, got: {joined[:500]}"
        assert "This is a generated test document for regression." in joined, (
            f"Expected source paragraph in XLSX, got: {joined[:500]}"
        )
    finally:
        workbook.close()

    assert len(list((settings.OUTPUT_DIR / conversion_id).glob("*"))) == 1


def test_docx_to_xlsx_output_is_real_ooxml_xlsx(tmp_path: Path):
    """Verify the output starts with the OOXML ZIP magic bytes (PK), proving it
    is a real XLSX container rather than a plain-text/fake file."""
    client = TestClient(app)

    docx_path = Path("tests/assets/regression/sample.docx")
    resp = client.post(
        "/convert",
        data={"target_format": "xlsx", "operation": "docx-to-xlsx"},
        files={
            "file": (
                docx_path.name,
                docx_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            )
        },
    )
    assert resp.status_code == 201, resp.text
    payload = resp.json()
    conversion_id, filename = Path(payload["download_path"].removeprefix("/download/")).parts
    local_path = settings.OUTPUT_DIR / conversion_id / filename

    with local_path.open("rb") as handle:
        header = handle.read(4)
    assert header == b"PK\x03\x04", f"Output is not a real XLSX (ZIP) file: {header!r}"
