"""
PROJECT: CONVERIGO
TEST SUITE: Certified PDF to XLSX Converter
STATUS: DEVELOPMENT (Not yet certified)

Basic test coverage for PDF→XLSX converter.
Ready for stability testing before certification.
"""

from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from app.core.settings import settings
from reportlab.pdfgen import canvas
from openpyxl import load_workbook

from app.main import app
from app.plugins.registry import registry


def create_simple_pdf() -> BytesIO:
    """Create a simple valid PDF."""
    buffer = BytesIO()
    pdf_canvas = canvas.Canvas(buffer)
    pdf_canvas.drawString(100, 750, "PDF to XLSX test")
    pdf_canvas.save()
    buffer.seek(0)
    return buffer


def test_pdf_to_excel_plugin_discovered():
    """TEST 001: Plugin is properly registered."""
    plugin = registry.get_plugin("pdf", "xlsx")
    assert plugin is not None
    assert plugin.slug == "pdf-to-excel"
    assert "xlsx" in plugin.target_formats


def test_pdf_to_excel_conversion_success():
    """TEST 002: PDF converts to XLSX successfully."""
    client = TestClient(app)
    simple_pdf = create_simple_pdf()
    
    response = client.post(
        "/convert",
        files={"file": ("test.pdf", simple_pdf, "application/pdf")},
        data={"target_format": "xlsx"},
    )
    
    assert response.status_code == 201
    assert response.json()["status"] == "success"


def test_pdf_to_excel_output_exists():
    """TEST 003: Output XLSX file is created."""
    client = TestClient(app)
    simple_pdf = create_simple_pdf()
    
    response = client.post(
        "/convert",
        files={"file": ("test.pdf", simple_pdf, "application/pdf")},
        data={"target_format": "xlsx"},
    )
    
    filename = response.json()["filename"]
    output_path = settings.OUTPUT_DIR / "document" / filename
    assert output_path.exists()
    
    output_path.unlink(missing_ok=True)


def test_pdf_to_excel_extension_correct():
    """TEST 004: Output file has correct XLSX extension."""
    client = TestClient(app)
    simple_pdf = create_simple_pdf()
    
    response = client.post(
        "/convert",
        files={"file": ("test.pdf", simple_pdf, "application/pdf")},
        data={"target_format": "xlsx"},
    )
    
    filename = response.json()["filename"]
    assert filename.endswith(".xlsx")
    
    output_path = settings.OUTPUT_DIR / "document" / filename
    assert output_path.suffix == ".xlsx"
    output_path.unlink(missing_ok=True)


def test_pdf_to_excel_not_corrupted():
    """TEST 005: Output XLSX file is valid and not corrupted."""
    client = TestClient(app)
    simple_pdf = create_simple_pdf()
    
    response = client.post(
        "/convert",
        files={"file": ("test.pdf", simple_pdf, "application/pdf")},
        data={"target_format": "xlsx"},
    )
    
    filename = response.json()["filename"]
    output_path = settings.OUTPUT_DIR / "document" / filename
    
    # File must be readable as valid XLSX
    try:
        workbook = load_workbook(str(output_path))
        assert workbook is not None
        assert len(workbook.sheetnames) > 0
    except Exception as e:
        raise AssertionError(f"Output XLSX is corrupted: {e}")
    finally:
        output_path.unlink(missing_ok=True)
